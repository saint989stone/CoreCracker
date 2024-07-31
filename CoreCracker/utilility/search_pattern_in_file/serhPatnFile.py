import openpyxl
import utilility.search_pattern_in_file.regular as regr

# def serhHost(strg):
#     hostList = ['error']
#     patnList = ['error']
#     lenhList = ['error']
#     for patn in regular.patnHost:  # перебираем регулярные выражения Host
#         host = re.search(patn, strg)  # осуществляем поиск строках значения соответствющим рег. выражениям Host
#         if host is not None:  # осуществляем проверку не пустых полученных значений
#             hostList.append(host[0])
#             patnList.append(patn)
#             lenhList.append(len(host[0]))
#
#     return hostList[-1], patnList[-1], lenhList[-1]

wb = openpyxl.load_workbook(filename='MPLS.xlsx')    #открываем фаил для чтения и записи
ws = wb.active  #получаем доступ к таблице на первом листе
cout = 1
for row in ws['B']:     #перебираем ячейки в столбце А
    coutStrg = str(cout)
    print(str(row.value))
    host, patn, lenh = regr.serhHostCEP(str(row.value))   #получаем значения host, patn, lenh
    ws['C' + coutStrg].value = host   #записываем значения в соответствующие ячейки
    ws['D' + coutStrg].value = patn
    ws['E' + coutStrg].value = lenh
    cout+=1
    print(cout)
wb.save(filename='MPLS.xlsx')   #записываем файл